from flask import Flask, request, jsonify, Response
import json
from flask_cors import CORS
import psycopg2
from psycopg2.extras import RealDictCursor
import os
from dotenv import load_dotenv
from openpyxl import Workbook
from io import BytesIO


load_dotenv()

app = Flask(__name__)
CORS(app)  # Habilitar CORS para todas las rutas

# Configuración de la base de datos

db_host = os.getenv("DB_HOST")
db_name = os.getenv("DB_NAME")
db_user = os.getenv("DB_USER")
db_password = os.getenv("DB_PASSWORD")


# Conexión a la base de datos
def get_db_connection():
    conn = psycopg2.connect(
        host=db_host, dbname=db_name, user=db_user, password=db_password
    )
    return conn


@app.route("/api/submit", methods=["POST"])
def submit_data():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)

        # Procesar datos del formulario
        form_data = request.form.to_dict()
        print("Datos del formulario:", form_data)

        # Procesar el archivo de la foto
        foto_funcionario = None
        if "fotoFuncionario" in request.files:
            file = request.files["fotoFuncionario"]
            if file.filename != "":
                foto_funcionario = file.read()

        # Insertar o recuperar responsable
        cursor.execute(
            "INSERT INTO responsables (name, phone) VALUES (%s, %s) RETURNING id",
            (form_data["responsableName"], form_data["responsablePhone"]),
        )
        responsable_id = cursor.fetchone()["id"]

        # Recuperar ID de la sede
        cursor.execute("SELECT id FROM sedes WHERE nombre = %s", (form_data["sede"],))
        sede = cursor.fetchone()
        if sede:
            sede_id = sede["id"]
        else:
            cursor.execute(
                "INSERT INTO sedes (nombre) VALUES (%s) RETURNING id",
                (form_data["sede"],),
            )
            sede_id = cursor.fetchone()["id"]

        # Insertar datos del funcionario
        cursor.execute(
            """
            INSERT INTO funcionarios (
                funcionario_id, name, age, gender, phone, tiene_carnet, razon_no_carnet, 
                foto, shirt_size, suit_size, shoe_size, num_carga_familiar, 
                num_fasdem_beneficiarios, sede_id, responsable_id, instagram, tiktok, facebook, cargo, tipo_personal, tipo_trabajador, adscripcion_nominal, ubicacion_fisica, funciones, estado_civil, nivel_academico, titulo_educacion_superior
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s, %s, %s, %s, %s, %s, %s,%s,%s,%s)
            RETURNING id
            """,
            (
                form_data["funcionarioID"],
                form_data["funcionarioName"],
                form_data["age"],
                form_data["gender"],
                form_data["funcionarioPhone"],
                form_data["tieneCarnet"] == "Sí",
                form_data.get("razonNoCarnet"),
                foto_funcionario,
                form_data.get("shirtSize"),
                form_data.get("suitSize"),
                form_data.get("shoeSize"),
                form_data.get("numCargaFamiliar"),
                form_data.get("numFasdemBeneficiarios"),
                sede_id,
                responsable_id,
                form_data.get("instagram"),
                form_data.get("tiktok"),
                form_data.get("facebook"),
                form_data.get("cargoActual"),
                form_data.get("tipoPersonal"),
                form_data.get("tipoTrabajador"),
                form_data.get("adscripcionNominal"),
                form_data.get("ubicacionFisica"),
                form_data.get("funcionesLaborales"),
                form_data.get("estadoCivil"),
                form_data.get("nivelAcademico"),
                form_data.get("tituloEducacionSuperior"),
            ),
        )
        funcionario_id = cursor.fetchone()["id"]

        cargas_familiares = json.loads(form_data.get("cargasFamiliares", "[]"))
        beneficiarios_fasdem = json.loads(form_data.get("beneficiariosFasdem", "[]"))
        hijos = json.loads(form_data.get("hijos", "[]"))

        for carga in cargas_familiares:
            cursor.execute(
                "INSERT INTO carga_familiar (funcionario_id, nombre, edad, patologias) VALUES (%s,%s,%s,%s)",
                [
                    funcionario_id,
                    carga.get("nombre"),
                    carga.get("edad"),
                    carga.get("patologias"),
                ],
            )

        for beneficiario in beneficiarios_fasdem:
            cursor.execute(
                "INSERT INTO beneficiarios_fasdem (funcionario_id, nombre, edad, patologias) VALUES (%s,%s,%s,%s)",
                [
                    funcionario_id,
                    beneficiario.get("nombre"),
                    beneficiario.get("edad"),
                    beneficiario.get("patologias"),
                ],
            )

        for hijo in hijos:
            cursor.execute(
                "INSERT INTO hijos (funcionario_id, nombre, edad, patologias) VALUES (%s,%s,%s,%s)",
                [
                    funcionario_id,
                    hijo.get("nombre"),
                    hijo.get("edad"),
                    hijo.get("patologias"),
                ],
            )

        conn.commit()
        return jsonify({"message": "Datos almacenados correctamente"}), 200

    except Exception as e:
        print("Error al procesar la solicitud:", e)
        return jsonify(
            {"error": "Error en el servidor, contacte al administrador"}
        ), 500

    finally:
        if conn:
            cursor.close()
            conn.close()


@app.route("/api/export_employees")
def export_employees():
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=RealDictCursor)

    # Obtener todas las sedes únicas
    cursor.execute("SELECT DISTINCT nombre FROM sedes")
    sedes = [row["nombre"] for row in cursor.fetchall()]

    # Lista de encabezados
    headers = [
        "Nombre Responsable",
        "Telefono Responsable",
        "Sede",
        "Cedula del Funcionario",
        "Nombre del Funcionario",
        "Edad del Funcionario",
        "Sexo del Funcionario",
        "Teléfono del Funcionario",
        "Cargo",
        "Tipo de Trabajador",
        "Tipo de Personal",
        "Adscripción Nominal",
        "Ubicación Física",
        "Funciones",
        "Posee carnet",
        "Estado Civil",
        "Nivel Académico",
        "Titulo Obtenido en Educación Superior",
        "Talla de Camisa",
        "Talla de Traje",
        "Talla de Zapatos",
        "Cantidad de Cargas Familiares",
        "Cantidad de Beneficiarios Inscritos en FASDEM",
        "Cantidad de Hijos",
        "Instagram",
        "TikTok",
        "Facebook",
    ]

    workbook = Workbook()

    for sede in sedes:
        # Crear una nueva hoja con el nombre de la sede
        sheet = workbook.create_sheet(
            title=sede[:31]
        )  # El nombre de una hoja no puede superar 31 caracteres

        # Escribir los encabezados en la fila 1
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        # Ejecutar la consulta para la sede actual
        cursor.execute(
            """
            SELECT 
                responsables.name AS nombre_responsable,
                responsables.phone AS telefono_responsable,
                sedes.nombre AS sede,
                func.funcionario_id,
                func.name as nombre_funcionario,
                func.age,
                func.gender,
                func.phone,
                func.cargo,
                func.tipo_trabajador,
                func.tipo_personal,
                func.adscripcion_nominal,
                func.ubicacion_fisica,
                func.funciones,
                func.tiene_carnet,
                func.estado_civil,
                func.nivel_academico,
                func.titulo_educacion_superior,
                func.shirt_size,
                func.suit_size,
                func.shoe_size,
                func.num_carga_familiar,
                func.num_fasdem_beneficiarios,
                func.instagram,
                func.tiktok,
                func.facebook,
                COUNT(hijos.id) AS num_hijos
            FROM funcionarios AS func
            INNER JOIN sedes
                ON sedes.id = func.sede_id
            INNER JOIN responsables
                ON responsables.id = func.responsable_id
            LEFT JOIN hijos
                ON hijos.funcionario_id = func.id
            WHERE sedes.nombre = %s
            GROUP BY 
                responsables.name,
                responsables.phone,
                sedes.nombre,
                func.funcionario_id,
                func.name,
                func.age,
                func.gender,
                func.phone,
                func.cargo,
                func.tipo_trabajador,
                func.tipo_personal,
                func.adscripcion_nominal,
                func.ubicacion_fisica,
                func.funciones,
                func.tiene_carnet,
                func.estado_civil,
                func.nivel_academico,
                func.titulo_educacion_superior,
                func.shirt_size,
                func.suit_size,
                func.shoe_size,
                func.num_carga_familiar,
                func.num_fasdem_beneficiarios,
                func.instagram,
                func.tiktok,
                func.facebook
            """,
            (sede,),
        )
        raw_data = cursor.fetchall()

        # Agregar los datos de la sede actual
        for data in raw_data:
            sheet.append(
                [
                    data["nombre_responsable"],
                    data["telefono_responsable"],
                    data["sede"],
                    data["funcionario_id"],
                    data["nombre_funcionario"],
                    data["age"],
                    data["gender"],
                    data["phone"],
                    data["cargo"],
                    data["tipo_trabajador"],
                    data["tipo_personal"],
                    data["adscripcion_nominal"],
                    data["ubicacion_fisica"],
                    data["funciones"],
                    "Si" if data["tiene_carnet"] else "No",
                    data["estado_civil"],
                    data["nivel_academico"],
                    "Ninguno"
                    if len(data["titulo_educacion_superior"]) < 1
                    else data["titulo_educacion_superior"],
                    data["shirt_size"],
                    data["suit_size"],
                    data["shoe_size"],
                    data["num_carga_familiar"],
                    data["num_fasdem_beneficiarios"],
                    data["num_hijos"],
                    data["instagram"],
                    data["tiktok"],
                    data["facebook"],
                ]
            )

        # Ajustar el ancho de las columnas automáticamente
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[
                0
            ].column_letter  # Obtiene la letra de la columna
            for cell in column_cells:
                try:
                    # Obtener la longitud del contenido de cada celda
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Ajuste extra para mayor claridad
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Eliminar la hoja predeterminada si no se usa
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Cerrar conexión
    cursor.close()
    conn.close()

    # Devolver el archivo como respuesta HTTP
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=puesto_a_puesto.xlsx"},
    )


if __name__ == "__main__":
    app.run(debug=True)
